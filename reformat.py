import openpyxl
import os

# Load the workbook
current_file_path = os.path.abspath(__file__)
workbook_path = os.path.join(os.path.dirname(current_file_path), '2023-11-19 Transaction Classifications.xlsx')
workbook = openpyxl.load_workbook(workbook_path)
sheet = workbook.active  # Assuming the relevant sheet is active or the first one

# Create a new sheet for the modified layout
if 'Modified Layout' in workbook.sheetnames:
    modified_sheet = workbook['Modified Layout']  # Use existing modified sheet
else:
    modified_sheet = workbook.create_sheet('Modified Layout')  # Create a new sheet

# Variables to track the current write positions for Level 1 and Level 2 categories
read_row = 1
write_row = 1

while read_row <= sheet.max_row:
    category_mark = sheet.cell(row=read_row, column=4).value
    category_name = sheet.cell(row=read_row, column=5).value
    category_description = sheet.cell(row=read_row + 1, column=5).value if read_row < sheet.max_row else None
    
    item_mark = sheet.cell(row=read_row, column=5).value
    item_name = sheet.cell(row=read_row, column=6).value
    item_description = sheet.cell(row=read_row + 1, column=6).value if read_row < sheet.max_row else None
    
    # Check for Category
    if category_mark == 'category':
        # Write Level 1 category name and category_description in the new layout
        modified_sheet.cell(row=write_row, column=1, value=category_name)
        modified_sheet.cell(row=write_row, column=3, value=category_description)
        write_row += 1  # Increment row for next entry

    # Check for Level 2 item
    elif item_mark =='item':
        # Write Level 2 category/item name and category_description in the new layout
        modified_sheet.cell(row=write_row, column=2, value=item_name)
        modified_sheet.cell(row=write_row, column=3, value=item_description)
        write_row += 1  # Increment row for next entry
    
    read_row += 2  # Increment row to skip the description row

# Save the workbook with modifications
modified_workbook_path = os.path.join(os.path.dirname(workbook_path), 'Modified 2023-11-19 Transaction Classifications.xlsx')  
workbook.save(modified_workbook_path)
print(f"Workbook has been modified and saved to {modified_workbook_path}")
