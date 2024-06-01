"""
This module loads an Excel workbook and finds all the formulas and 
replaces them with python equivalents. 
"""
import os
import re
from pprint import pprint
import openpyxl
from openpyxl import load_workbook
import pandas as pd


def load_workbook_from_file(filename):
    """
    Load an Excel workbook from a file using openpyxl.

    Parameters:
    filename (str): The name of the Excel file.

    Returns:
    Workbook: The loaded workbook.
    """
    file_path = os.path.join(os.getcwd(), filename)
    workbook = load_workbook(file_path)
    return workbook

def load_df_from_file(filename, sheet_name):
    """
    Load an Excel worksheet into a pandas dataframe.

    Parameters:
    filename (str): The name of the Excel file.
    sheet_name (str): The name of the worksheet.

    Returns:
    DataFrame: The loaded dataframe.
    """
    file_path = os.path.join(os.getcwd(), filename)
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    return df

def get_cell_value_from_df(df, cell_ref):
    """
    Get the value of a cell in a dataframe.

    Parameters:
    df (DataFrame): The dataframe.
    row (int): The row index.
    col (str): The column name.

    Returns:
    The value of the cell.
    """

    # Split the cell reference into row and column parts
    col, row = re.findall(r'[A-Z]+|\d+', cell_ref)
    # convert the column name to a zero-based index
    col = ord(col.upper()) - ord('A')      
    # Convert the row index to a zero-based index
    row = int(row) - 1  
    return df.loc[row, col]

def get_formulas_from_sheet(workbook, sheet_name):
    """
    Get all the formulas in a sheet along with their cell locations.
    """
    sheet = workbook[sheet_name]
    extracted_formulas = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str) \
            and cell.value.startswith('='):
                extracted_formulas[cell.coordinate] = cell.value
    return extracted_formulas

def resolve_excel_cell_reference(sheet, df, cell_reference):
    """
    Resolves an Excel cell reference and returns the value of the cell.

    Parameters:
    sheet (Worksheet): The Excel worksheet.
    cell_reference (str): The cell reference.

    Returns:
    The value of the cell.
    """
    # Strip any inverted commas and dollar characters from the cell reference
    cell_reference = cell_reference.replace("'", "").replace("$", "")

    # Get the cell value
    # cell_value = sheet[cell_reference].value
    # get the cell value from the pandas dataframe
    cell_value = get_cell_value_from_df(df, cell_reference)

    return cell_value



def get_cell_value(sheet, cell_reference):
    # Attempt to get the value directly from the cell
    
    try:
        # Try to convert it to an integer in case it's a direct number
        cell_value = sheet[cell_reference].value
        return int(cell_value)
    except ValueError:
        # If it's not a direct number, process it as an expression
        return evaluate_expression(sheet, cell_value)

def evaluate_expression(sheet, expression):
    # Safe evaluation context: Only allow the operators needed
    allowed_operators = {"+": (lambda x, y: x + y)}
    
    # Split the expression assuming a format like "I5 + 7"
    parts = expression.split()
    cell_ref = parts[0]
    operator = parts[1]
    number = int(parts[2])
    
    # Get the cell value for the cell reference
    referenced_value = get_cell_value(sheet, cell_ref)
    
    # Evaluate using the allowed operators
    if operator in allowed_operators:
        return allowed_operators[operator](referenced_value, number)
    else:
        raise ValueError("Unsupported operation")




def convert_vlookup_to_named_references(this_sheet,formula):
    # TODO combine this with the get_lookup_sheet_and_col_name function
    """ 
    Extract the sheet name and column index from a VLOOKUP formula.
        
    Parameters:
        worksheet_name (str): The Excel worksheet.
        input_formulas (dict): A dictionary of cell references and their formulas.

    Returns:
        the sheet name and column index
    """

    # extract the VLOOKUP formula from within the larger formula
    # e.g. =IF(ANALYSYS!$C3="RTT",0,VLOOKUP(ANALYSYS!$A3,'MW DB'!A:AAI,151,0))
    # extract the VLOOKUP formula VLOOKUP(ANALYSYS!$A3,'MW DB'!A:AAI,151,0)
    match = re.search(r'VLOOKUP\([^)]+\)', formula)
    formula = match.group(0) if match else None
    try:
        # Split the formula to isolate the parts
        parts = formula.split(',')
        # Extract the worksheet name between the single quotes
        extracted_sheet_name = re.search(r"'(.+?)'", parts[1]).group(1)
        # Extract the VLOOKUP column number directly from the second part
        extracted_col_index = parts[2]
        if not extracted_col_index.isdigit():
            # If the column number is a cell reference, resolve it to an absolute number
            # Strip any inverted commas and dollar characters from the cell reference
            extracted_col_index = extracted_col_index.replace("'", "").replace("$", "")
            # extract_col_index = get_cell_value(this_sheet, extracted_col_index)
            # Get the cell value
            # extracted_col_index = this_sheet[extracted_col_index].value
        else:
            extracted_col_index = int(extracted_col_index)
        return extracted_sheet_name, extracted_col_index
    except IndexError:
        print("Parsing error: Not enough parts in the formula. Please check the formula format.")
    except AttributeError:
        print("Parsing error: No valid worksheet name found. Please check the formula format.")
    except ValueError:
        print("Parsing error: Column index is not a number. Please check the formula format.")
    return None, None

def get_lookup_sheet_and_col_name(formula, ws_with_formula, df, workbook):
    """
    Extracts the sheet name and column name from a VLOOKUP formula.

    Parameters:
        formula (str): The VLOOKUP formula.
        ws_with_formula (str): The the Excel worksheet where the formula is located.
        wookbook (Workbook): The Excel workbook that contains the referenced ws_with_formula.

    Returns:
    tuple: The lookup sheet name and column name.
    """
    lookup_sheet_name, col_index = convert_vlookup_to_named_references(ws_with_formula,formula)
    lookup_sheet = workbook[lookup_sheet_name]
    # resolve the column index to and absolute number
    if isinstance(col_index, str):
        col_index = resolve_excel_cell_reference(ws_with_formula, df, col_index)
        if isinstance(col_index, str):
            # the referenced cell is also and expression
            col_index =get_cell_value(ws_with_formula, col_index)
    column_name = get_col_name_from_ws_col_index( lookup_sheet, col_index)
    return lookup_sheet_name, column_name


def get_col_name_from_ws_col_index( worksheet, col_index):
    """ Get the column name from a worksheet based on the column index.
        Assume that the column name is the first string value found 
        in the first three rows of the column.
    """

    # Find the column title in the first three rows
    column_name = None
    for row in worksheet.iter_rows(min_row=1, max_row=3, min_col=col_index, max_col=col_index):
        for cell in row:
            # ignore cells that contain formulas
            if isinstance(cell.value, str) and not cell.value.startswith('='):
                column_name = cell.value
                break
        if column_name:
            break

    return column_name

def extract_vlookup_formula(expression):
    """ Extract the VLOOKUP formula from an input expression."""
    # Define a regular expression to match the VLOOKUP pattern
    vlookup_pattern = r"VLOOKUP\(.*?\)"

    # Search for the pattern in the input expression
    match = re.search(vlookup_pattern, expression)

    # If a match is found, return the VLOOKUP formula, otherwise return None
    if match:
        return match.group(0)
    return None

def create_column_name_dictionary(worksheet):
    """ Create a dictionary mapping column letters to column names."""

    # Dictionary to store column letters to column name mappings
    col_names = {}

    # Access the second row in the worksheet
    col_name_row = worksheet[2]

    # Iterate through each cell in the second row
    for cell in col_name_row:
        if cell.value is not None:
            # Get the column letter from the cell coordinate
            column_letter = cell.column_letter
            # Map the column letter to the cell value (column name)
            col_names[column_letter] = cell.value

    return col_names

# def replace_dependent_excel_cell_ref_w_df_index(
#         input_sheet_name,
#         input_column_names,
#         input_formulas):
#     """
#     Replace the dependent variable cell references in the formulas with a dataframe index.
#     dataframe name and column are the is the sheet name and column name

#     Parameters:
#         input sheet name (str): The Excel worksheet. Uses as the dataframe name
#         Input column names (dict): dictionary of column names in the excel sheet
#         input_formulas (dict): A dictionary of cell references and their formulas.

#     Returns:
#         A modified dictionary of formulas with cell references replaced by dataframe index.
#     """

#     updated_formulas = {}
#     for cell_ref, formula in input_formulas.items():
#         column_letter = cell_ref[0]  # Assuming cell reference like 'C3'
#         column_name = input_column_names.get(column_letter, 'UnknownColumn')
#         # TODO replace the data index.  the dataframe name is the input sheet name
#         # TODO The data index is the column name
#         updated_formulas[column_name] = formula
#     return updated_formulas

def replace_excel_cell_refs_with_df_index(worksheet_name, input_formulas):
    """ 
    Replace all occurrences of cell references
    with a dataframe index where dataframe name is the sheet name in the formula
    or default worksheet name if not found.  
    dataframe column names are the same as the excel column names
        
    Parameters:
        worksheet_name (str): The Excel worksheet.
        input_formulas (dict): A dictionary of cell references and their formulas.

    Returns:
        A modified dictionary of formulas with cell references replaced by dataframe index.
    """
    renamed_formulas = {}

    # identify cell references on the same sheet
    for cell_ref, formula in input_formulas.items():
        # returns a list of tuples with the worksheet name (or None)  and the cell reference
        cell_references = re.findall(r'(\w+!)?([A-Z]+\d+)', formula)

        # replace the cell references with the dataframe index
        for ref in cell_references:
            sheet_name, cell_ref = ref
            col, row = re.findall(r'[A-Z]+|\d+', cell_ref)
            if sheet_name:  # if a sheet name was found
                sheet_name = sheet_name[:-1]  # remove the trailing "!"
            else:  # if no sheet name was found
                sheet_name = worksheet_name  # use the default worksheet name
            df_index = f"{sheet_name}.loc[{int(row)-1}, '{col}']"
            formula = formula.replace(cell_ref, df_index)


            renamed_formulas[cell_ref] = formula

    return renamed_formulas

def replace_vlookup_with_df_index(this_worksheet, input_formulas,df, workbook):
    """ 
    Replace the VLOOKUP formulas with a dataframe index.
    Assume dataframe name and column name is the same as 
    the sheet name and the column name in the excel sheet.
        
    Parameters:
        this_worksheet (str): The Excel worksheet.
        input_formulas (dict): A dictionary of cell references and their formulas.

    Returns:
        A modified dictionary of formulas with cell references replaced by dataframe index.
    """
    revised_formulas = {}
    for cell_ref, formula in input_formulas.items():
        print(f"cell_ref: {cell_ref}, formula: {formula}")
        if 'VLOOKUP' in formula:
            lookup_sheet_name, col_name = get_lookup_sheet_and_col_name(
                                            formula, this_worksheet,df, workbook)
            df_index = f"{lookup_sheet_name}.loc[{col_name}]"
            # Extract the full VLOOKUP formula and replace it with the df_index
            vlookup_formula = extract_vlookup_formula(formula)
            formula = formula.replace(vlookup_formula, df_index)
        revised_formulas[cell_ref] = formula
    return revised_formulas

# the name of the sheet to analyze
FILE_NAME = 'OQ_sample_TMDB.xlsx'
SHEET_NAME = 'ANALYSYS'

# Load the workbook and worksheet using openpyxl to get the formulas (not the values)
wb = load_workbook_from_file(FILE_NAME)
analysis_sheet = wb[SHEET_NAME]
# load the data into a pandas dataframe to get the values of the cells (not formulas)
analysis_df = load_df_from_file(FILE_NAME, SHEET_NAME)



# Get all the column names and formulas from the worksheet
column_names = create_column_name_dictionary(analysis_sheet)
formulas = get_formulas_from_sheet(wb, 'ANALYSYS')

# Show the found formulas and their locations
print("\n" * 3)
print("Original formulas in the excel:")
pprint(formulas)
print("\n" * 3)
print("Column names:")
pprint(column_names)
print("\n" * 3)

# Replace the VLOOKUP formulas with the dataframe index
updated_formulas = replace_vlookup_with_df_index(analysis_sheet, formulas,analysis_df, wb)
print("\n" * 3)
print("Updated formulas after replacing VLOOKUPS:")
pprint(updated_formulas)

# Replace the dependent cell references in the formulas with a dataframe index
# updated_formulas = replace_dependent_excel_cell_ref_w_df_index(SHEET_NAME, column_names, formulas)
# pprint(updated_formulas)

# replace the  cell references in the formulas with a dataframe index
updated_formulas = replace_excel_cell_refs_with_df_index(SHEET_NAME, formulas)

print("\n" * 3)
print("Updated formulas after replacing cell refs:")
pprint(updated_formulas)


# TODO clean up the formulas

# TODO iterate over all rows in the analysis sheet

# TODO generate test file
